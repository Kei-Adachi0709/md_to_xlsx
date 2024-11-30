import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

def load_markdown_file(md_file):
    """MarkdownファイルからJSONデータ部分を読み込む"""
    with open(md_file, 'r', encoding='utf-8') as file:
        md_text = file.read()
        json_start = md_text.find('[')
        json_end = md_text.rfind(']') + 1
        json_text = md_text[json_start:json_end]
        return json.loads(json_text)

def apply_style(cell, style, logged_warnings=None):
    """セルにスタイルを適用する"""
    if not style:
        return

    # ログ済みの警告を追跡
    if logged_warnings is None:
        logged_warnings = set()

    # フォントスタイル
    font_kwargs = {}
    if 'italic' in style.get('font', {}):
        font_kwargs['italic'] = style['font']['italic']
    if 'bold' in style.get('font', {}):
        font_kwargs['bold'] = style['font']['bold']
    if 'size' in style.get('font', {}):
        font_kwargs['size'] = style['font']['size']
    if 'name' in style.get('font', {}):
        font_kwargs['name'] = style['font']['name']

    # 色の設定 (RGB → aRGB変換)
    if 'color' in style:
        color = style['color']
        if len(color) == 7 and color.startswith('#'):  # e.g., #RRGGBB
            color = f"FF{color[1:]}"  # Convert to FFRRGGBB
        font_kwargs['color'] = color

    if font_kwargs:
        try:
            cell.font = Font(**font_kwargs)
        except ValueError as e:
            if font_kwargs.get('color') not in logged_warnings:
                print(f"Warning: Invalid font color '{font_kwargs.get('color')}'.")
                logged_warnings.add(font_kwargs.get('color'))

    if 'bgcolor' in style:
        bgcolor = style['bgcolor']
        if len(bgcolor) == 7 and bgcolor.startswith('#'):  # e.g., #RRGGBB
            bgcolor = f"FF{bgcolor[1:]}"  # Convert to FFRRGGBB
        cell.fill = PatternFill(start_color=bgcolor, end_color=bgcolor, fill_type='solid')

    # 罫線 (RGB → aRGB変換)
    if 'border' in style:
        border_sides = {}
        for side, border_data in style['border'].items():
            border_color = border_data[1]
            if len(border_color) == 7 and border_color.startswith('#'):  # e.g., #RRGGBB
                border_color = f"FF{border_color[1:]}"  # Convert to FFRRGGBB

            try:
                border_sides[side] = Side(style=border_data[0], color=border_color)
            except ValueError as e:
                if border_color not in logged_warnings:
                    print(f"Warning: Invalid border color '{border_color}'.")
                    logged_warnings.add(border_color)
                continue
        cell.border = Border(**border_sides)

    # アンダーライン
    if style.get('underline', False):
        cell.font = Font(underline="single")

    # 打消し線
    if 'strike' in style:
        cell.font = Font(strike=style['strike'])


def create_excel_sheet(wb, sheet_data, logged_warnings=None):
    """Excelのワークブックにシートを追加し、JSONデータを設定する"""
    if logged_warnings is None:
        logged_warnings = set()
    sheet_name = sheet_data.get('name', 'Sheet1')
    ws = wb.create_sheet(title=sheet_name)
    rows = sheet_data.get('rows', {})

    for row_idx, row_data in rows.items():
        if row_idx == 'len':
            continue

        row_idx = int(row_idx)
        cells = row_data.get('cells', {})

        for col_idx_str, cell_data in cells.items():
            col_idx = int(col_idx_str)
            cell_value = cell_data.get('text', '')
            cell = ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
            style = sheet_data.get('styles', [])[cell_data.get('style', -1)]
            apply_style(cell, style, logged_warnings)

def main():
    # コマンドライン引数からMarkdownファイル名を取得
    if len(sys.argv) != 2:
        print("Usage: md_to_xlsx.py <markdown_file>")
        sys.exit(1)

    md_file = sys.argv[1]
    data = load_markdown_file(md_file)

    # Excelワークブックの作成
    wb = Workbook()

    for sheet in data:
        create_excel_sheet(wb, sheet)

    # デフォルトのシートを削除
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # 出力ファイル名を決定
    output_file = md_file.replace('.md', '.xlsx')

    # Excelファイルに保存
    wb.save(output_file)
    print(f"Excelファイル '{output_file}' が生成されました。")

if __name__ == "__main__":
    main()
